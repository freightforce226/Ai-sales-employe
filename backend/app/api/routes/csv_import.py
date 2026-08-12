from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
import io
import csv
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from app.db.session import get_db_session
from app.core.auth import get_current_user
from app.models.user import User
from app.core.config import get_settings
from app.core.logging import get_logger
from pydantic import BaseModel, Field
from typing import Dict, Any, Optional, List
import httpx
import uuid
import json

router = APIRouter(prefix="/api/v1/import", tags=["CSV Import"])
settings = get_settings()
logger = get_logger(__name__)

class StartImportRequest(BaseModel):
    storage_path: str
    column_mapping: Dict[str, str]
    headers: List[str]
    header_row: int = Field(0, ge=0, le=20)
    file_name: Optional[str] = "import.csv"

def parse_uploaded_file(file_content: bytes, filename: str, sheet_name: Optional[str] = None) -> tuple[List[List[str]], List[str]]:
    """
    Parses any uploaded spreadsheet/text file in-memory using capability-based matching.
    Returns a tuple (rows, sheet_names).
    """
    MAX_FILE_SIZE = 50 * 1024 * 1024
    if len(file_content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File exceeds maximum allowed size of 50MB."
        )

    # Try XLSX / XLSM (openpyxl)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet_names = wb.sheetnames
        rows = []
        if sheet_name and sheet_name in sheet_names:
            sheet = wb[sheet_name]
            for r in sheet.iter_rows(values_only=True):
                rows.append([str(val) if val is not None else '' for val in r])
        else:
            # Detect automatically
            max_non_empty_cells = -1
            for sheet in wb.worksheets:
                filled_cells = 0
                temp_rows = []
                for r in sheet.iter_rows(values_only=True):
                    if any(val is not None and str(val).strip() != '' for val in r):
                        filled_cells += sum(1 for val in r if val is not None and str(val).strip() != '')
                        temp_rows.append([str(val) if val is not None else '' for val in r])
                if filled_cells > max_non_empty_cells:
                    max_non_empty_cells = filled_cells
                    rows = temp_rows
        return rows, sheet_names
    except Exception:
        pass

    # Try XLS (xlrd)
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_content)
        sheet_names = wb.sheet_names()
        rows = []
        if sheet_name and sheet_name in sheet_names:
            sheet = wb.sheet_by_name(sheet_name)
            for r_idx in range(sheet.nrows):
                r = sheet.row_values(r_idx)
                rows.append([str(val) if val is not None else '' for val in r])
        else:
            # Detect automatically
            max_non_empty_cells = -1
            for sheet_idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_idx)
                filled_cells = 0
                temp_rows = []
                for r_idx in range(sheet.nrows):
                    r = sheet.row_values(r_idx)
                    if any(val is not None and str(val).strip() != '' for val in r):
                        filled_cells += sum(1 for val in r if val is not None and str(val).strip() != '')
                        temp_rows.append([str(val) if val is not None else '' for val in r])
                if filled_cells > max_non_empty_cells:
                    max_non_empty_cells = filled_cells
                    rows = temp_rows
        return rows, sheet_names
    except Exception:
        pass

    # Try parsing as text-based (CSV / TSV / TXT)
    text = None
    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            text = file_content.decode(enc)
            break
        except UnicodeDecodeError:
            continue

    if text is not None:
        delimiter = ','
        sample = text[:2000]
        if '\t' in sample:
            if filename.lower().endswith('.tsv') or sample.count('\t') > sample.count(','):
                delimiter = '\t'
        try:
            f = io.StringIO(text)
            reader = csv.reader(f, delimiter=delimiter)
            rows = list(reader)
            return rows, []
        except Exception:
            pass

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="The uploaded file format is unsupported, corrupted, or could not be parsed."
    )

def score_row(cells: list[str]) -> int:
    clean_cells = [str(c).strip().lower() for c in cells if str(c).strip() != '']
    if len(clean_cells) <= 1:
        return 0
        
    confidence_keywords = [
        'company', 'importer', 'contact', 'email', 'mail', 'phone', 'address', 
        'industry', 'sector', 'website', 'linkedin', 'name', 's/l', 'serial', 
        'goods', 'description', 'detail', 'client', 'phone number', 'zip', 
        'state', 'city', 'country', 'visit', 'remarks'
    ]
    
    ignore_phrases = [
        'importers list', 'customer report', 'export data', 'report list', 
        'export list', 'export report', 'import list', 'import report',
        'customer visit report'
    ]
    
    score = 0
    row_text = ' '.join(clean_cells)
    for phrase in ignore_phrases:
        if phrase in row_text:
            score -= 50
            
    for cell in clean_cells:
        if len(cell) > 40:
            score -= 15
            continue
        for kw in confidence_keywords:
            if cell == kw:
                score += 25
            elif kw in cell:
                score += 10
        score += 1
    return max(0, score)

def detect_header_row(rows: list[list[str]]) -> int:
    best_index = 0
    max_score = -1
    limit = min(15, len(rows))
    for i in range(limit):
        score = score_row(rows[i])
        if score > max_score:
            max_score = score
            best_index = i
    return best_index

@router.post("/upload")
async def upload_csv(
    file: UploadFile = File(...),
    header_row: int = Form(0),
    sheet_name: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user)
):
    """
    Parses any supported file, normalizes it to a clean CSV in-memory,
    uploads it to Supabase Storage, and logs telemetry.
    """
    import time
    start_time = time.time()

    file_content = await file.read()
    
    # 1. Parse File
    parse_start = time.time()
    rows, sheet_names = parse_uploaded_file(file_content, file.filename, sheet_name=sheet_name)
    parse_time = time.time() - parse_start

    # Auto-detect header row if input is 0
    detected_header_row = header_row
    if header_row == 0:
        detected_header_row = detect_header_row(rows)

    if detected_header_row >= len(rows):
        detected_header_row = 0

    # Validate header row is not empty
    header = rows[detected_header_row]
    if not any(str(c).strip() for c in header):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Detected header row is empty"
        )

    # 2. Normalization
    norm_start = time.time()
    raw_normalized_rows = rows[detected_header_row:]
    normalized_rows = [
        r for r in raw_normalized_rows
        if any(str(c).strip() for c in r)
    ]

    if not normalized_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File contains no data after normalization"
        )

    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\n")
    writer.writerows(normalized_rows)
    normalized_content = out.getvalue().encode("utf-8")
    norm_time = time.time() - norm_start

    # 3. Upload to storage
    upload_start = time.time()
    org_id = current_user.organization_id
    file_id = uuid.uuid4()
    
    clean_filename = file.filename
    allowed_exts = ('.csv', '.tsv', '.txt', '.xlsx', '.xls', '.xlsm')
    for ext in allowed_exts:
        if clean_filename.lower().endswith(ext):
            clean_filename = clean_filename[:-len(ext)]
            break
    storage_filename = f"{clean_filename}.csv"
    storage_path = f"{org_id}/{file_id}_{storage_filename}"

    import urllib.parse
    safe_storage_path = urllib.parse.quote(storage_path)
    supabase_upload_url = f"{settings.supabase_url}/storage/v1/object/csv-imports/{safe_storage_path}"
    
    async with httpx.AsyncClient() as client:
        try:
            res = await client.post(
                supabase_upload_url,
                content=normalized_content,
                headers={
                    "apikey": settings.supabase_service_role_key,
                    "Authorization": f"Bearer {settings.supabase_service_role_key}",
                    "Content-Type": "text/csv"
                }
            )
            
            if res.status_code not in (200, 201):
                raise HTTPException(
                    status_code=status.HTTP_502_BAD_GATEWAY,
                    detail=f"Supabase Storage upload failed: {res.text}"
                )
            upload_time = time.time() - upload_start
            total_time = time.time() - start_time

            logger.info(
                "Telemetry: File upload parsing completed",
                file_name=file.filename,
                rows_count=len(normalized_rows),
                parse_time_ms=int(parse_time * 1000),
                normalization_time_ms=int(norm_time * 1000),
                upload_time_ms=int(upload_time * 1000),
                total_time_ms=int(total_time * 1000)
            )
            
            client_headers = [str(h).strip() for h in header]
            all_rows_preview = rows[:15]
            return {
                "storage_path": storage_path, 
                "file_name": file.filename,
                "header_row_used": detected_header_row,
                "headers": client_headers,
                "all_rows_preview": all_rows_preview,
                "sheet_names": sheet_names
            }
        except Exception as e:
            if isinstance(e, HTTPException):
                raise
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Upload request error: {str(e)}"
            )

@router.post("/start")
async def start_import(
    request: StartImportRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """
    Validates, parses, normalizes, and extracts data from the uploaded file in-memory.
    Writes the standardized canonical CSV to a NEW storage path, updates the batch path,
    deletes the original file, and triggers the n8n webhook (Fail Fast Policy).
    """
    org_id = current_user.organization_id
    batch_id = uuid.uuid4()

    # Pre-create import batch record in processing state
    try:
        await db.execute(
            text("""
                INSERT INTO import_batches (id, organization_id, status, file_name, file_path, header_row_used, successful_rows, failed_rows, total_rows, error_log)
                VALUES (:id, :org_id, 'processing', :file_name, :file_path, :header_row, 0, 0, 0, '[]')
            """),
            {
                "id": batch_id,
                "org_id": org_id,
                "file_name": request.file_name,
                "file_path": request.storage_path,
                "header_row": 0
            }
        )
        await db.commit()
    except Exception as init_err:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to initialize import batch: {str(init_err)}"
        )

    import re
    email_pattern = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')
    phone_pattern = re.compile(r'(?:\+?\d{1,3}[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}')

    def extract_contact_info(text: str) -> dict:
        if not text:
            return {}
        text_str = str(text).strip()
        
        emails = email_pattern.findall(text_str)
        phones = phone_pattern.findall(text_str)
        
        websites = []
        for word in text_str.split():
            word_clean = word.strip().lower()
            if 'www.' in word_clean or 'http://' in word_clean or 'https://' in word_clean:
                websites.append(word)
                
        email = emails[0] if emails else None
        phone = phones[0] if phones else None
        website = websites[0] if websites else None
        
        parts = [p.strip() for p in re.split(r'[\n\r|,|&|/]', text_str) if p.strip()]
        
        contact_name = None
        designation = None
        company_name = None
        
        for p in parts:
            if email and email in p:
                continue
            if phone and phone in p:
                continue
            if website and website in p:
                continue
                
            p_lower = p.lower()
            if any(title in p_lower for title in ('manager', 'director', 'ceo', 'vp', 'executive', 'sales', 'rep', 'lead', 'partner')):
                designation = p
                continue
                
            if not contact_name:
                if any(c.isalpha() for c in p) and len(p) < 40:
                    contact_name = p
            elif not company_name:
                if any(c.isalpha() for c in p) and len(p) < 80:
                    company_name = p
                    
        return {
            "email": email,
            "phone": phone,
            "website": website,
            "contact_name": contact_name,
            "designation": designation,
            "company_name": company_name
        }

    new_storage_path = request.storage_path.replace(".csv", "_normalized.csv")

    try:
        # 1. Download
        import urllib.parse
        safe_storage_path = urllib.parse.quote(request.storage_path)
        supabase_storage_url = f"{settings.supabase_url}/storage/v1/object/csv-imports/{safe_storage_path}"
        async with httpx.AsyncClient() as client:
            res_down = await client.get(
                supabase_storage_url,
                headers={
                    "apikey": settings.supabase_service_role_key,
                    "Authorization": f"Bearer {settings.supabase_service_role_key}"
                }
            )
            if res_down.status_code != 200:
                raise ValueError(f"Failed to retrieve uploaded file from Supabase Storage: {res_down.text}")
            csv_content = res_down.text

        # 2. Parse & Extract
        f = io.StringIO(csv_content)
        reader = csv.reader(f)
        rows = list(reader)
        if not rows:
            raise ValueError("Downloaded CSV is empty")
            
        original_headers = [h.strip() for h in rows[0]]
        data_rows = rows[1:]

        # Get mapped column indices
        col_indices = {}
        for db_field in ('company_name', 'contact_name', 'contact_email', 'industry'):
            original_header = request.column_mapping.get(db_field)
            if original_header and original_header in original_headers:
                col_indices[db_field] = original_headers.index(original_header)
            else:
                col_indices[db_field] = None

        normalized_data = [["company_name", "contact_name", "contact_email", "industry"]]
        
        for row in data_rows:
            if not row:
                continue
                
            raw_company_cell = row[col_indices["company_name"]] if col_indices["company_name"] is not None and col_indices["company_name"] < len(row) else ""
            raw_name_cell = row[col_indices["contact_name"]] if col_indices["contact_name"] is not None and col_indices["contact_name"] < len(row) else ""
            raw_email_cell = row[col_indices["contact_email"]] if col_indices["contact_email"] is not None and col_indices["contact_email"] < len(row) else ""
            raw_industry_cell = row[col_indices["industry"]] if col_indices["industry"] is not None and col_indices["industry"] < len(row) else ""

            extracted_company = extract_contact_info(raw_company_cell)
            extracted_name = extract_contact_info(raw_name_cell)
            extracted_email = extract_contact_info(raw_email_cell)

            # Email mapping
            row_contact_email = raw_email_cell.strip()
            if not email_pattern.match(row_contact_email) and extracted_email.get("email"):
                row_contact_email = extracted_email["email"]
            if not row_contact_email and extracted_company.get("email"):
                row_contact_email = extracted_company["email"]
            if not row_contact_email and extracted_name.get("email"):
                row_contact_email = extracted_name["email"]

            # Company name mapping
            row_company_name = raw_company_cell.strip()
            if not row_company_name and extracted_email.get("company_name"):
                row_company_name = extracted_email["company_name"]
            if not row_company_name and extracted_name.get("company_name"):
                row_company_name = extracted_name["company_name"]

            # Contact name mapping
            row_contact_name = raw_name_cell.strip()
            if not row_contact_name and extracted_email.get("contact_name"):
                row_contact_name = extracted_email["contact_name"]
            if not row_contact_name and extracted_company.get("contact_name"):
                row_contact_name = extracted_company["contact_name"]

            row_industry = raw_industry_cell.strip()

            normalized_data.append([row_company_name, row_contact_name, row_contact_email, row_industry])

        # 3. Canonical Header Validation
        expected_headers = ["company_name", "contact_name", "contact_email", "industry"]
        actual_headers = normalized_data[0]
        if actual_headers != expected_headers:
            raise ValueError(f"Canonical header mismatch. Expected {expected_headers}, got {actual_headers}")

        # 4. Generate standard CSV & Upload to NEW storage path
        out = io.StringIO()
        writer = csv.writer(out, lineterminator="\n")
        writer.writerows(normalized_data)
        normalized_content = out.getvalue().encode("utf-8")

        safe_new_storage_path = urllib.parse.quote(new_storage_path)
        supabase_upload_url = f"{settings.supabase_url}/storage/v1/object/csv-imports/{safe_new_storage_path}"

        async with httpx.AsyncClient() as client:
            res_up = await client.post(
                supabase_upload_url,
                content=normalized_content,
                headers={
                    "apikey": settings.supabase_service_role_key,
                    "Authorization": f"Bearer {settings.supabase_service_role_key}",
                    "Content-Type": "text/csv"
                }
            )
            if res_up.status_code not in (200, 201):
                raise ValueError(f"Supabase Storage upload of normalized CSV failed: {res_up.text}")

        # 5. Update database batch with NEW path
        await db.execute(
            text("UPDATE import_batches SET file_path = :file_path WHERE id = :id"),
            {"file_path": new_storage_path, "id": batch_id}
        )
        await db.commit()

        # 6. Delete old file from storage
        supabase_delete_url = f"{settings.supabase_url}/storage/v1/object/csv-imports/{safe_storage_path}"
        async with httpx.AsyncClient() as client:
            await client.delete(
                supabase_delete_url,
                headers={
                    "apikey": settings.supabase_service_role_key,
                    "Authorization": f"Bearer {settings.supabase_service_role_key}"
                }
            )

        # 7. Persist Successful import mapping
        try:
            mapping_res = await db.execute(
                text("SELECT id, headers FROM import_mappings WHERE organization_id = :org_id"),
                {"org_id": org_id}
            )
            existing_mappings = mapping_res.fetchall()
            matched_mapping_id = None
            target_headers_set = set(request.headers)
            for row in existing_mappings:
                row_id, row_headers = row
                try:
                    loaded_headers = json.loads(row_headers) if isinstance(row_headers, str) else row_headers
                except Exception:
                    loaded_headers = row_headers
                if isinstance(loaded_headers, list) and set(loaded_headers) == target_headers_set:
                    matched_mapping_id = row_id
                    break
            
            mapping_name = f"Template for {request.file_name}"
            if matched_mapping_id:
                await db.execute(
                    text("UPDATE import_mappings SET column_mapping = :column_mapping, updated_at = NOW(), mapping_name = :name WHERE id = :id"),
                    {"id": matched_mapping_id, "column_mapping": json.dumps(request.column_mapping), "name": mapping_name}
                )
            else:
                await db.execute(
                    text("INSERT INTO import_mappings (id, organization_id, mapping_name, headers, column_mapping) VALUES (:id, :org_id, :name, :headers, :column_mapping)"),
                    {"id": uuid.uuid4(), "org_id": org_id, "name": mapping_name, "headers": json.dumps(request.headers), "column_mapping": json.dumps(request.column_mapping)}
                )
            await db.commit()
        except Exception as mapping_err:
            logger.error("Failed to persist successful import mapping template", error=str(mapping_err))
            await db.rollback()

    except Exception as norm_err:
        await db.rollback()
        # Mark batch as failed
        await db.execute(
            text("""
                UPDATE import_batches 
                SET status = 'failed', error_log = :error_log, completed_at = NOW() 
                WHERE id = :id
            """),
            {
                "id": batch_id,
                "error_log": json.dumps([{"error": f"Normalization Failure: {str(norm_err)}"}])
            }
        )
        await db.commit()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Import normalization failed: {str(norm_err)}"
        )

    # 8. Trigger n8n Webhook (Only after successful verification & path update)
    n8n_webhook_url = settings.n8n_webhook_url
    payload = {
        "import_batch_id": str(batch_id),
        "organization_id": str(org_id),
        "storage_path": new_storage_path,
        "header_row": 0,
        "column_mapping": {
            "company_name": "company_name",
            "contact_name": "contact_name",
            "contact_email": "contact_email",
            "industry": "industry"
        }
    }
    headers = {
        "X-API-Key": settings.n8n_service_api_key
    }
    
    logger.info(
        "Triggering n8n Webhook Workflow 1",
        webhook_url=n8n_webhook_url,
        request_body=payload,
        request_headers=headers
    )
    
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                n8n_webhook_url,
                json=payload,
                headers=headers,
                timeout=10.0
            )
            logger.info(
                "n8n Webhook response received",
                status_code=response.status_code,
                response_body=response.text
            )
            
            if response.status_code < 200 or response.status_code >= 300:
                raise ValueError(f"n8n Webhook returned status code {response.status_code}: {response.text}")
        except Exception as webhook_err:
            logger.error("n8n webhook execution failed", error=str(webhook_err))
            await db.execute(
                text("""
                    UPDATE import_batches 
                    SET status = 'failed', error_log = :error_log, completed_at = NOW() 
                    WHERE id = :id
                """),
                {
                    "id": batch_id,
                    "error_log": json.dumps([{"error": f"Webhook trigger failure: {str(webhook_err)}"}])
                }
            )
            await db.commit()
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail=f"Failed to trigger import orchestration: {str(webhook_err)}"
            )

    return {"batch_id": batch_id, "status": "processing"}

@router.get("/batches/{id}")
async def get_batch_status(
    id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """
    Returns the real-time statistics and status of the selected batch.
    """
    try:
        res = await db.execute(
            text("""
                SELECT status, file_name, successful_rows, failed_rows, total_rows, created_at
                FROM import_batches
                WHERE id = :id AND organization_id = :org_id
            """),
            {"id": id, "org_id": current_user.organization_id}
        )
        row = res.fetchone()
        if not row:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Import batch not found."
            )
        
        successful = row[2] or 0
        failed = row[3] or 0
        total = row[4] or 0
        return {
            "id": id,
            "status": row[0],
            "file_name": row[1],
            "success_count": successful,
            "duplicate_count": 0,
            "error_count": failed,
            "processed_rows": successful + failed,
            "total_rows": total,
            "created_at": str(row[5])
        }
    except Exception as e:
        if isinstance(e, HTTPException):
            raise
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch batch progress: {str(e)}"
        )

@router.get("/batches/{id}/errors")
async def get_batch_errors(
    id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """
    Exposes error details logged by n8n during validation.
    """
    try:
        res = await db.execute(
            text("SELECT error_log FROM import_batches WHERE id = :id AND organization_id = :org_id"),
            {"id": id, "org_id": current_user.organization_id}
        )
        row = res.fetchone()
        if not row:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Import batch not found."
            )
        
        errors = []
        if row[0]:
            try:
                errors = json.loads(row[0]) if isinstance(row[0], str) else row[0]
            except Exception:
                errors = []
        return {"errors": errors}
    except Exception as e:
        if isinstance(e, HTTPException):
            raise
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch batch error logs: {str(e)}"
        )

@router.get("/history")
async def get_import_history(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """
    Lists past batches processed for the organization.
    """
    try:
        res = await db.execute(
            text("""
                SELECT id, file_name, status, successful_rows, failed_rows, total_rows, created_at
                FROM import_batches
                WHERE organization_id = :org_id
                ORDER BY created_at DESC
            """),
            {"org_id": current_user.organization_id}
        )
        history = []
        for r in res.fetchall():
            successful = r[3] or 0
            failed = r[4] or 0
            history.append({
                "id": str(r[0]),
                "file_name": r[1],
                "status": r[2],
                "success_count": successful,
                "duplicate_count": 0,
                "error_count": failed,
                "total_rows": r[5] or 0,
                "created_at": str(r[6])
            })
        return history
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch history logs: {str(e)}"
        )

@router.get("/mappings")
async def get_saved_mappings(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session)
):
    """
    Fetches all saved column mappings/templates for the organization.
    """
    try:
        res = await db.execute(
            text("""
                SELECT id, mapping_name, headers, column_mapping, created_at
                FROM import_mappings
                WHERE organization_id = :org_id
                ORDER BY updated_at DESC
            """),
            {"org_id": current_user.organization_id}
        )
        mappings = []
        for r in res.fetchall():
            mappings.append({
                "id": str(r[0]),
                "mapping_name": r[1],
                "headers": r[2],
                "column_mapping": r[3],
                "created_at": str(r[4])
            })
        return mappings
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch saved mappings: {str(e)}"
        )
